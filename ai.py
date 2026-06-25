#!/usr/bin/env python3
"""
MI AI - Complete Python Backend v2.0
Multi-Key Rotation + Auto Fallback + Enhanced PDF
By Muaaz Iqbal | Muslim Islam Org
بِسْمِ اللَّهِ الرَّحْمَٰنِ الرَّحِيمِ
"""

import os
import json
import time
import base64
import zipfile
import tempfile
import requests
import io
import re
import threading
from pathlib import Path
from datetime import datetime
from flask import Flask, request, jsonify, send_file, Response, stream_with_context
from flask_cors import CORS
from groq import Groq

# ─────────────────────────────────────────
# INIT
# ─────────────────────────────────────────
app = Flask(__name__)
CORS(app, origins="*")

UPLOAD_DIR = Path(tempfile.gettempdir()) / "mi_ai_uploads"
UPLOAD_DIR.mkdir(exist_ok=True)

# ─────────────────────────────────────────
# MULTI-KEY ROTATION
# ─────────────────────────────────────────
GROQ_API_KEYS = [
    os.getenv("GROQ_API_KEY_1", "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"),
    os.getenv("GROQ_API_KEY_2", "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"),
    os.getenv("GROQ_API_KEY_3", "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"),
    os.getenv("GROQ_API_KEY_4", "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"),
    os.getenv("GROQ_API_KEY_5", "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"),
    os.getenv("GROQ_API_KEY_6", "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"),
]
GROQ_API_KEYS = [k for k in GROQ_API_KEYS if k]  # Remove empty

key_lock = threading.Lock()
_current_key_index = 0
_key_fail_counts = {}

def get_current_key():
    return GROQ_API_KEYS[_current_key_index]

def rotate_key(failed_idx=None):
    global _current_key_index, _key_fail_counts
    with key_lock:
        if failed_idx is not None:
            _key_fail_counts[failed_idx] = _key_fail_counts.get(failed_idx, 0) + 1
        for i in range(1, len(GROQ_API_KEYS) + 1):
            next_idx = (_current_key_index + i) % len(GROQ_API_KEYS)
            if _key_fail_counts.get(next_idx, 0) < 2:
                _current_key_index = next_idx
                print(f"[MI AI] Switched to API key #{_current_key_index + 1}")
                return True
        # All exhausted — reset
        _key_fail_counts = {}
        _current_key_index = 0
        print("[MI AI] All keys cycled, resetting")
        return False

def get_groq_client():
    """Get Groq client with current key."""
    return Groq(api_key=get_current_key())

def groq_call_with_rotation(messages, model="llama-3.3-70b-versatile", max_tokens=8192, temperature=0.7):
    """Call Groq API with automatic key rotation on failure."""
    last_err = None
    for attempt in range(len(GROQ_API_KEYS)):
        key_idx = _current_key_index
        try:
            client = get_groq_client()
            resp = client.chat.completions.create(
                model=model,
                messages=messages,
                max_tokens=max_tokens,
                temperature=temperature,
            )
            _key_fail_counts[key_idx] = 0  # Reset on success
            return resp.choices[0].message.content or ""
        except Exception as e:
            err_str = str(e).lower()
            last_err = e
            print(f"[MI AI] Key #{key_idx+1} error: {e}")
            if "rate_limit" in err_str or "429" in err_str or "401" in err_str or "403" in err_str:
                rotate_key(key_idx)
                time.sleep(0.3)
                continue
            # For other errors, still rotate and try
            rotate_key(key_idx)
            time.sleep(0.2)
    raise Exception(f"All {len(GROQ_API_KEYS)} API keys failed. Last error: {last_err}")

def groq_stream_with_rotation(messages, model="llama-3.1-8b-instant", max_tokens=8192, temperature=0.7):
    """Stream Groq API with automatic key rotation on failure."""
    last_err = None
    for attempt in range(len(GROQ_API_KEYS)):
        key_idx = _current_key_index
        try:
            client = get_groq_client()
            stream = client.chat.completions.create(
                model=model,
                messages=messages,
                max_tokens=max_tokens,
                temperature=temperature,
                stream=True,
            )
            _key_fail_counts[key_idx] = 0
            for chunk in stream:
                delta = chunk.choices[0].delta.content or ""
                if delta:
                    yield delta
            return  # Success
        except Exception as e:
            err_str = str(e).lower()
            last_err = e
            print(f"[MI AI] Stream key #{key_idx+1} error: {e}")
            if "rate_limit" in err_str or "429" in err_str or "401" in err_str:
                rotate_key(key_idx)
                time.sleep(0.3)
                continue
            yield f"\n\n⚠️ Error: {e}"
            return
    yield f"\n\n❌ All API keys exhausted. Please wait and try again."

# ─────────────────────────────────────────
# GROQ MODELS
# ─────────────────────────────────────────
MODELS = {
    "llama-3.3-70b-versatile":    {"name": "Llama 3.3 70B",        "type": "fast",     "ctx": 128000},
    "llama-3.1-8b-instant":       {"name": "Llama 3.1 8B Instant",  "type": "instant",  "ctx": 128000},
    "meta-llama/llama-4-scout-17b-16e-instruct":         {"name": "Mixtral 8x7B",          "type": "balanced", "ctx": 32768},
    "llama-3.1-8b-instant":               {"name": "Gemma 2 9B",            "type": "smart",    "ctx": 8192},
    "deepseek-r1-distill-llama-70b": {"name": "DeepSeek R1 Pro",   "type": "thinking", "ctx": 128000},
    "deepseek-r1-distill-qwen-32b":  {"name": "DeepSeek R1 Qwen",  "type": "thinking", "ctx": 128000},
    "qwen-qwq-32b":               {"name": "QwQ 32B Reasoning",     "type": "thinking", "ctx": 128000},
    "meta-llama/llama-4-maverick-17b-128e-instruct": {"name":"Llama 4 Maverick","type":"latest","ctx":128000},
    "meta-llama/llama-4-scout-17b-16e-instruct":     {"name":"Llama 4 Scout",   "type":"latest","ctx":128000},
    "compound-beta":              {"name": "Compound Beta",         "type": "multi",    "ctx": 128000},
    "compound-beta-mini":         {"name": "Compound Beta Mini",    "type": "multi",    "ctx": 128000},
    "llama3-groq-70b-8192-tool-use-preview": {"name":"Llama 70B Tools","type":"tool","ctx":8192},
}

# ─────────────────────────────────────────
# SYSTEM PROMPTS
# ─────────────────────────────────────────
SYSTEMS = {
    "chat": """You are MI AI — Advanced Intelligence by Muaaz Iqbal (Muslim Islam Org).
Brilliant, helpful, comprehensive. Use rich markdown. Write COMPLETE answers, never truncate.
For code: full working implementation always — 2000-5000+ lines when asked. No shortcuts.
Begin Islamic answers with بِسْمِ اللَّهِ""",

    "pro": """You are MI AI Pro Thinking Mode by Muaaz Iqbal (Muslim Islam Org).
Think deeply step-by-step. Show reasoning process. Extremely detailed answers.
Code: COMPLETE full implementations — all functions, all edge cases, all error handling.""",

    "code": """You are MI AI Code Expert by Muaaz Iqbal (Muslim Islam Org) — world's best programmer.
ALWAYS write COMPLETE working code. NEVER use '...' or truncate. Write 2000-5000+ lines when asked.
Include: full error handling, comprehensive comments, tests, documentation.
Expert in: Python, JS, TS, React, Node, Go, Rust, C++, Java, PHP, Flutter, SQL, etc.""",

    "files": """You are MI AI File Analysis Expert by Muaaz Iqbal (Muslim Islam Org).
Analyze: PDFs, images, ZIPs, code, CSV, JSON, Excel, Word, M3U, any format.
Provide: comprehensive analysis, key insights, patterns, recommendations, statistics.""",

    "pdf": """You are MI AI Book & PDF Generator by Muaaz Iqbal (Muslim Islam Org).
Create complete books — full chapters, detailed paragraphs, 500+ pages when asked.
Professional: intro, chapters, sub-sections, conclusion, references.
Islamic books: include Quran verses, hadith, proper Arabic.""",

    "quran": """You are MI AI Islamic Knowledge by Muaaz Iqbal (Muslim Islam Org).
بِسْمِ اللَّهِ الرَّحْمَٰنِ الرَّحِيمِ
Expert: Quran (114 surahs, full tafseer), Hadith (Bukhari, Muslim, Tirmidhi etc),
Fiqh (all 4 madhabs), Islamic history, duas, prayer, Arabic.
Format: Arabic text → transliteration → translation → explanation.""",

    "web": """You are MI AI Web Research Expert by Muaaz Iqbal (Muslim Islam Org).
Give comprehensive info on any topic. Cite sources. Multiple perspectives. Latest known info.""",
}

# ─────────────────────────────────────────
# TELEGRAM BOT
# ─────────────────────────────────────────
TELEGRAM_TOKEN = os.getenv("TELEGRAM_TOKEN", "")
TELEGRAM_ADMIN_CHAT_ID = os.getenv("TELEGRAM_ADMIN_CHAT_ID", "")
TELEGRAM_API = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}"

def telegram_send(chat_id, text, parse_mode="Markdown"):
    """Send message via Telegram."""
    if not TELEGRAM_TOKEN:
        return
    try:
        requests.post(f"{TELEGRAM_API}/sendMessage", json={
            "chat_id": chat_id,
            "text": text[:4096],  # Telegram limit
            "parse_mode": parse_mode
        }, timeout=10)
    except Exception as e:
        print(f"[Telegram] Send error: {e}")

def telegram_notify_admin(message):
    """Notify admin via Telegram."""
    if TELEGRAM_ADMIN_CHAT_ID:
        telegram_send(TELEGRAM_ADMIN_CHAT_ID, message)

@app.route("/api/telegram/webhook", methods=["POST"])
def telegram_webhook():
    """Handle Telegram bot messages."""
    data = request.json or {}
    message = data.get("message", {})
    chat_id = message.get("chat", {}).get("id")
    text = message.get("text", "").strip()
    username = message.get("from", {}).get("username", "User")

    if not chat_id or not text:
        return jsonify({"ok": True})

    # Commands
    if text == "/start":
        telegram_send(chat_id,
            "🤖 *MI AI — Powered by Muaaz Iqbal | Muslim Islam Org*\n\n"
            "بِسْمِ اللَّهِ الرَّحْمَٰنِ الرَّحِيمِ\n\n"
            "Mujhe kuch bhi poochein! Main aapki madad karoonga.\n\n"
            "Commands:\n/start - Start\n/help - Help\n/models - Available models"
        )
        return jsonify({"ok": True})

    if text == "/help":
        telegram_send(chat_id,
            "📖 *MI AI Help*\n\n"
            "• Koi bhi sawaal poochein\n"
            "• Code likhwayein\n"
            "• Islamic knowledge\n"
            "• File analysis (coming soon)\n\n"
            "Just type your question!"
        )
        return jsonify({"ok": True})

    if text == "/models":
        model_list = "\n".join([f"• {v['name']}" for v in list(MODELS.values())[:8]])
        telegram_send(chat_id, f"🧠 *Available Models:*\n\n{model_list}")
        return jsonify({"ok": True})

    # Regular chat
    try:
        telegram_send(chat_id, "⏳ Sooch raha hoon...")
        reply = groq_call_with_rotation([
            {"role": "system", "content": SYSTEMS["chat"]},
            {"role": "user", "content": text}
        ], model="llama-3.3-70b-versatile", max_tokens=2048)
        
        # Truncate for Telegram if needed
        if len(reply) > 4000:
            reply = reply[:3900] + "\n\n... _(continued — please ask for more)_"
        
        telegram_send(chat_id, reply)
        
        # Notify admin of new user
        if TELEGRAM_ADMIN_CHAT_ID and str(chat_id) != str(TELEGRAM_ADMIN_CHAT_ID):
            telegram_notify_admin(f"👤 *New message from @{username}*\n\n{text[:200]}")

    except Exception as e:
        telegram_send(chat_id, f"❌ Error: {str(e)}\n\nPlease try again.")

    return jsonify({"ok": True})

@app.route("/api/telegram/set-webhook", methods=["POST"])
def set_telegram_webhook():
    """Set Telegram webhook URL."""
    data = request.json or {}
    webhook_url = data.get("url", "")
    if not webhook_url or not TELEGRAM_TOKEN:
        return jsonify({"error": "Missing webhook URL or token"}), 400
    
    try:
        resp = requests.post(f"{TELEGRAM_API}/setWebhook", json={"url": webhook_url}, timeout=10)
        return jsonify(resp.json())
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ─────────────────────────────────────────
# ROUTES
# ─────────────────────────────────────────

@app.route("/")
def root():
    return jsonify({
        "status": "MI AI Backend Running",
        "version": "2.0",
        "by": "Muaaz Iqbal | Muslim Islam Org",
        "keys_loaded": len(GROQ_API_KEYS),
        "active_key": f"#{_current_key_index + 1}",
        "telegram": "configured" if TELEGRAM_TOKEN else "not configured"
    })

@app.route("/api/models")
def get_models():
    return jsonify({"models": MODELS})

@app.route("/api/keys/status")
def keys_status():
    """Check which API key is active."""
    return jsonify({
        "total_keys": len(GROQ_API_KEYS),
        "active_key_index": _current_key_index + 1,
        "fail_counts": {str(k+1): v for k, v in _key_fail_counts.items()}
    })

@app.route("/api/chat", methods=["POST"])
def chat():
    data = request.json or {}
    user_msg    = data.get("message", "")
    mode        = data.get("mode", "chat")
    model       = data.get("model", "llama-3.3-70b-versatile")
    history     = data.get("history", [])
    stream_mode = data.get("stream", False)

    sys_prompt = SYSTEMS.get(mode, SYSTEMS["chat"])
    messages = [{"role": "system", "content": sys_prompt}]
    messages += history[-16:]
    messages.append({"role": "user", "content": user_msg})

    if stream_mode:
        def generate():
            for chunk in groq_stream_with_rotation(messages, model):
                yield f"data: {json.dumps({'delta': chunk})}\n\n"
            yield "data: [DONE]\n\n"
        return Response(
            stream_with_context(generate()),
            mimetype="text/event-stream",
            headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"}
        )

    try:
        reply = groq_call_with_rotation(messages, model)
        return jsonify({"reply": reply, "model": model, "mode": mode})
    except Exception as e:
        return jsonify({"error": str(e), "reply": f"❌ Error: {str(e)}"}), 500

@app.route("/api/analyze-file", methods=["POST"])
def analyze_file():
    """Analyze uploaded files: PDF, images, ZIP, CSV, JSON, code, etc."""
    if "file" not in request.files:
        return jsonify({"error": "No file provided"}), 400

    f = request.files["file"]
    fname = f.filename or "file"
    ext = fname.rsplit(".", 1)[-1].lower() if "." in fname else "bin"
    model = request.form.get("model", "llama-3.3-70b-versatile")

    content = ""
    analysis_type = "general"

    try:
        raw = f.read()
        TEXT_EXTS = {"txt","md","log","yaml","yml","toml","ini","cfg","conf","html","xml","csv","tsv",
                     "json","jsonl","js","ts","py","php","java","cpp","c","cs","go","rs","rb","swift",
                     "kt","dart","sql","sh","bash","jsx","tsx","vue","r","scala","lua","m3u","m3u8"}
        CODE_EXTS = {"js","ts","py","php","java","cpp","c","cs","go","rs","rb","swift","kt","dart",
                     "sql","sh","bash","jsx","tsx","vue","r","scala","lua"}

        if ext in TEXT_EXTS:
            try:
                content = raw.decode("utf-8", errors="replace")
                analysis_type = "code" if ext in CODE_EXTS else "text"
            except:
                content = "[Binary content — cannot decode as text]"

        elif ext == "zip":
            analysis_type = "zip"
            try:
                with zipfile.ZipFile(io.BytesIO(raw)) as z:
                    names = z.namelist()
                    content = f"ZIP Archive: {fname}\nTotal files: {len(names)}\n\nContents:\n"
                    for n in names[:50]:
                        content += f"  {'[DIR]' if n.endswith('/') else '[FILE]'} {n}\n"
                    content += "\n--- File previews ---\n"
                    count = 0
                    for n in names:
                        if count >= 8: break
                        nx = n.rsplit(".", 1)[-1].lower() if "." in n else ""
                        if nx in TEXT_EXTS and not n.endswith("/"):
                            try:
                                fc = z.read(n).decode("utf-8", errors="replace")
                                content += f"\n=== {n} ({len(fc.splitlines())} lines) ===\n{fc[:2000]}\n"
                                count += 1
                            except:
                                pass
            except Exception as ze:
                content = f"ZIP read error: {ze}"

        elif ext == "pdf":
            analysis_type = "pdf"
            try:
                import pypdf
                reader = pypdf.PdfReader(io.BytesIO(raw))
                pages_text = []
                for i, page in enumerate(reader.pages[:20]):
                    text = page.extract_text() or ""
                    if text.strip():
                        pages_text.append(f"[Page {i+1}]\n{text}")
                content = "\n\n".join(pages_text) if pages_text else "[PDF has no extractable text]"
                content = f"PDF: {fname} ({len(reader.pages)} pages)\n\n" + content[:8000]
            except Exception as pe:
                content = f"PDF read error: {pe}"

        elif ext in {"png","jpg","jpeg","gif","webp","bmp"}:
            analysis_type = "image"
            b64 = base64.b64encode(raw).decode()
            content = f"[Image file: {fname}, size: {len(raw)} bytes, base64 length: {len(b64)}]"

        elif ext in {"xlsx","xls","ods"}:
            analysis_type = "spreadsheet"
            try:
                import openpyxl
                wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True)
                content = f"Excel: {fname}\nSheets: {', '.join(wb.sheetnames)}\n\n"
                for sheet_name in wb.sheetnames[:3]:
                    ws = wb[sheet_name]
                    content += f"\n=== Sheet: {sheet_name} ===\n"
                    rows = []
                    for i, row in enumerate(ws.iter_rows(values_only=True)):
                        if i >= 50: break
                        rows.append("\t".join(str(c) if c is not None else "" for c in row))
                    content += "\n".join(rows)
            except Exception as xe:
                content = f"Excel read error: {xe}"

        else:
            try:
                content = raw.decode("utf-8", errors="replace")[:3000]
            except:
                content = f"[Binary file: {fname}, {len(raw)} bytes]"

        sys_prompt = SYSTEMS.get("files", SYSTEMS["chat"])
        prompt = f"""Analyze this {analysis_type} file named "{fname}":

{content[:6000]}

Provide:
1. **Summary & Overview** — what is this file about
2. **Key Information** — important data, functions, content extracted
3. **Structure Analysis** — how it's organized
4. **Observations** — patterns, issues, notable things
5. **Recommendations** — improvements, fixes, suggestions
6. **Technical Details** — size, format specifics"""

        analysis = groq_call_with_rotation(
            [{"role": "system", "content": sys_prompt}, {"role": "user", "content": prompt}],
            model=model, max_tokens=4096
        )
        return jsonify({"analysis": analysis, "file": fname, "type": analysis_type, "size": len(raw)})

    except Exception as e:
        return jsonify({"error": str(e), "analysis": f"❌ Analysis failed: {e}"}), 500

@app.route("/api/generate-pdf", methods=["POST"])
def generate_pdf_content():
    """Generate complete book/document content for PDF creation."""
    data    = request.json or {}
    topic   = data.get("topic", "")
    pages   = int(data.get("pages", 50))
    type_   = data.get("type", "general")
    details = data.get("details", "")
    lang    = data.get("language", "English")

    if not topic:
        return jsonify({"error": "Topic required"}), 400

    # Outline
    outline_prompt = f"""Create a detailed book outline for: "{topic}"
Type: {type_} | Language: {lang} | Target pages: {pages}
{details}
Include: Title, Preface, 10-15 chapters with sub-sections, Conclusion, Index.
Format clearly."""

    outline = groq_call_with_rotation(
        [{"role": "system", "content": SYSTEMS["pdf"]},
         {"role": "user", "content": outline_prompt}],
        model="llama-3.3-70b-versatile", max_tokens=3000
    )

    # Chapters
    chapters_needed = max(5, min(15, pages // 10))
    chapters = []
    for i in range(1, chapters_needed + 1):
        words_per_chapter = max(800, (pages * 300) // chapters_needed)
        chapter_prompt = f"""Write Chapter {i} for a book about "{topic}".
Type: {type_} | Language: {lang}
Outline reference: {outline[:600]}
Write exactly {words_per_chapter}+ words. Be detailed, informative, engaging.
Include: chapter intro, 3-5 detailed sections, examples/cases, chapter summary.
{'Include relevant Quranic verses and hadith with Arabic text.' if type_ == 'islamic' else ''}
{'Write in Urdu language.' if lang == 'Urdu' else ''}"""

        chapter_content = groq_call_with_rotation(
            [{"role": "system", "content": SYSTEMS["pdf"]},
             {"role": "user", "content": chapter_prompt}],
            model="llama-3.3-70b-versatile", max_tokens=4096
        )
        chapters.append({"chapter": i, "content": chapter_content})

    return jsonify({
        "topic": topic,
        "outline": outline,
        "chapters": chapters,
        "type": type_,
        "pages_target": pages
    })

@app.route("/api/generate-code", methods=["POST"])
def generate_code():
    data  = request.json or {}
    desc  = data.get("description", "")
    lang  = data.get("language", "Python")
    model = data.get("model", "llama-3.3-70b-versatile")

    prompt = f"""Write a COMPLETE, PRODUCTION-READY {lang} implementation for:
{desc}

REQUIREMENTS:
- Write FULL code — no truncation, no "..." placeholders
- Include all imports, functions, classes
- Full error handling and input validation
- Comprehensive comments explaining logic
- Example usage at the bottom
- At least 200+ lines for complex requests"""

    try:
        code = groq_call_with_rotation(
            [{"role": "system", "content": SYSTEMS["code"]},
             {"role": "user", "content": prompt}],
            model=model, max_tokens=8192
        )
        return jsonify({"code": code, "language": lang})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/fix-code", methods=["POST"])
def fix_code():
    data  = request.json or {}
    code  = data.get("code", "")
    error = data.get("error", "")
    lang  = data.get("language", "auto-detect")
    model = data.get("model", "llama-3.3-70b-versatile")

    prompt = f"""Fix this {lang} code completely:

```{lang}
{code}
```
{f'Error message: {error}' if error else ''}

Return:
1. **Root Cause** — what exactly is wrong
2. **Fixed Code** — complete corrected version (ALL lines, nothing truncated)
3. **Changes Made** — list each fix with explanation
4. **Prevention** — how to avoid this issue"""

    try:
        fix = groq_call_with_rotation(
            [{"role": "system", "content": SYSTEMS["code"]},
             {"role": "user", "content": prompt}],
            model=model, max_tokens=8192
        )
        return jsonify({"fix": fix, "language": lang})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/generate-quiz", methods=["POST"])
def generate_quiz():
    data   = request.json or {}
    topic  = data.get("topic", "")
    num    = int(data.get("questions", 20))
    level  = data.get("level", "medium")
    model  = data.get("model", "llama-3.3-70b-versatile")

    resp = groq_call_with_rotation([
        {"role": "system", "content": "You are an expert educator. Create comprehensive quizzes."},
        {"role": "user", "content": f"""Create a {level} difficulty quiz about "{topic}" with {num} questions.
Mix: multiple choice (A/B/C/D), true/false, and short answer.
Format each question clearly with ANSWER key at the end.
Return as valid JSON: {{"quiz":[{{"q":"question","type":"mcq/tf/short","options":["A","B","C","D"],"answer":"correct"}}]}}"""}
    ], model, max_tokens=4096)

    try:
        m = re.search(r"\{[\s\S]*\}", resp)
        quiz = json.loads(m.group()) if m else {"raw": resp}
    except:
        quiz = {"raw": resp}
    return jsonify(quiz)

@app.route("/api/code-review", methods=["POST"])
def code_review():
    data  = request.json or {}
    code  = data.get("code", "")
    lang  = data.get("language", "auto-detect")
    model = data.get("model", "llama-3.3-70b-versatile")

    resp = groq_call_with_rotation([
        {"role": "system", "content": SYSTEMS["code"]},
        {"role": "user", "content": f"""Review this {lang} code and provide:
1. **Issues Found** — bugs, security issues, performance problems
2. **Fixed Code** — complete corrected version (ALL lines)
3. **Improvements** — optimizations, best practices
4. **Explanation** — what each fix does and why

Code to review:
```{lang}
{code}
```"""}
    ], model, max_tokens=8192)
    return jsonify({"review": resp})

@app.route("/api/translate", methods=["POST"])
def translate():
    data   = request.json or {}
    text   = data.get("text", "")
    target = data.get("target_lang", "Arabic")
    model  = data.get("model", "llama-3.1-8b-instant")

    resp = groq_call_with_rotation([
        {"role": "user", "content": f"Translate to {target}. Return ONLY the translation:\n\n{text}"}
    ], model, max_tokens=4096)
    return jsonify({"translation": resp, "target": target})

@app.route("/api/summarize", methods=["POST"])
def summarize():
    data  = request.json or {}
    text  = data.get("text", "")
    style = data.get("style", "comprehensive")
    model = data.get("model", "llama-3.3-70b-versatile")

    resp = groq_call_with_rotation([
        {"role": "user", "content": f"""Provide a {style} summary. Include: main points, key findings, important details, conclusion.

Text: {text[:8000]}"""}
    ], model, max_tokens=2048)
    return jsonify({"summary": resp})

@app.route("/api/islamic", methods=["POST"])
def islamic_knowledge():
    """Islamic knowledge endpoint."""
    data  = request.json or {}
    query = data.get("query", "")
    model = data.get("model", "llama-3.3-70b-versatile")

    resp = groq_call_with_rotation([
        {"role": "system", "content": SYSTEMS["quran"]},
        {"role": "user", "content": query}
    ], model, max_tokens=4096)
    return jsonify({"answer": resp, "bismillah": "بِسْمِ اللَّهِ الرَّحْمَٰنِ الرَّحِيمِ"})

@app.route("/api/health")
def health():
    return jsonify({
        "status": "ok",
        "time": datetime.now().isoformat(),
        "by": "Muaaz Iqbal | Muslim Islam Org",
        "api_keys": len(GROQ_API_KEYS),
        "active_key": f"#{_current_key_index + 1}"
    })

# ─────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────
def _md_to_html(text):
    text = re.sub(r"^# (.+)$", r"<h1>\1</h1>", text, flags=re.MULTILINE)
    text = re.sub(r"^## (.+)$", r"<h2>\1</h2>", text, flags=re.MULTILINE)
    text = re.sub(r"^### (.+)$", r"<h3>\1</h3>", text, flags=re.MULTILINE)
    text = re.sub(r"\*\*(.+?)\*\*", r"<strong>\1</strong>", text)
    text = re.sub(r"\*(.+?)\*", r"<em>\1</em>", text)
    text = re.sub(r"^- (.+)$", r"<li>\1</li>", text, flags=re.MULTILINE)
    text = text.replace("\n\n", "</p><p>")
    return "<p>" + text + "</p>"

# ─────────────────────────────────────────
if __name__ == "__main__":
    port = int(os.getenv("PORT", 5000))
    print(f"MI AI Backend v2.0 starting on port {port}")
    print(f"API Keys loaded: {len(GROQ_API_KEYS)}")
    print("By Muaaz Iqbal | Muslim Islam Org")
    print("بِسْمِ اللَّهِ الرَّحْمَٰنِ الرَّحِيمِ")
    app.run(host="0.0.0.0", port=port, debug=False)
